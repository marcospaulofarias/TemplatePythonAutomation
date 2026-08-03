from loguru import logger
import openpyxl # type: ignore
from resources.FilesManager import FilesManager
from resources.PrintAutomation import PrintAutomation

class Xlsx:
    """Classe para manipulação de arquivos Excel (.xlsx) usando openpyxl.

    :param process_id: Identificador do processo para logs e rastreamento.
    :param process_type: Tipo do processo para logs e rastreamento.
    :param process_machine: Nome da máquina para logs e rastreamento.
    :returns None: se ocorrer nenhuma falha
    """
    def __init__(self, process_id: str, process_type: str, process_machine: str) -> None:
        logger.debug(f"Xlsx.__init__: process_id={process_id} process_type={process_type} process_machine={process_machine}")
        self.workbook = None
        self.filesmanager = FilesManager(process_id=process_id, process_type=process_type, process_machine=process_machine)
        self.printautomation = PrintAutomation(process_id=process_id, process_type=process_type, process_machine=process_machine)

    def _create_xlsx(self, xlsx_file: str) -> None:
        logger.debug(f"Xlsx._create_xlsx: xlsx_file={xlsx_file}")
        try:
            if not self.filesmanager.verify_exists_file(xlsx_file):
                self.workbook = openpyxl.Workbook()
                self.save_workbook(xlsx_file=xlsx_file)
                logger.debug(f"Xlsx._create_xlsx: created workbook {xlsx_file}")
        except Exception as error_x:
            logger.critical(f'O arquivo "{xlsx_file}" não pôde ser criado.\nErro: {error_x}.')
            raise RuntimeError(f'O arquivo "{xlsx_file}" não pôde ser criado.\nErro: {error_x}.') from error_x

    def _get_workbook(self, xlsx_file: str, create_xlsx: bool = False) -> openpyxl.Workbook:
        logger.debug(f"Xlsx._get_workbook: xlsx_file={xlsx_file} create_xlsx={create_xlsx}")
        try:
            if create_xlsx:
                self._create_xlsx(xlsx_file=xlsx_file)
            self.workbook = openpyxl.load_workbook(filename=xlsx_file)
            logger.debug(f"Xlsx._get_workbook: loaded workbook {xlsx_file}")
            return self.workbook
        except Exception as error_x:
            logger.critical(f'Erro ao capturar arquivo xlsx "{xlsx_file}".\nErro: {error_x}')
            raise RuntimeError(f'Erro ao capturar arquivo xlsx "{xlsx_file}".\nErro: {error_x}') from error_x

    def get_sheet_names(self, xlsx_file: str) -> list[str]:
        """Retorna uma lista com os nomes das planilhas no arquivo Excel.
        
        :return: lista com os nomes das planilhas no arquivo.
        :raises RunTimeError: se não puder retornar os nomes das planilhas.
        """
        logger.debug(f"Xlsx.get_sheet_names: xlsx_file={xlsx_file}")
        try:
            self.workbook = self._get_workbook(xlsx_file)
            sheetnames = self.workbook.sheetnames
            logger.debug(f"Xlsx.get_sheet_names: sheetnames={sheetnames}")
            return sheetnames
        except Exception as error_x:
            logger.critical(f'Não foi possível retornar as planilhas do arquivo "{xlsx_file}".\nErro: {error_x}')
            raise RuntimeError(f'Não foi possível retornar as planilhas do arquivo "{xlsx_file}".') from error_x
    
    def get_sheet(self, xlsx_file: str, sheet_name: str, create_sheet: bool = False) -> openpyxl.workbook:
        """Retorna a planilha especificada pelo nome.

        :param sheet_name: nome da planilha a ser retornada.
        :param create_sheet: criar planilha se não existir?
        :return: objeto da planilha.
        :raises RunTimeError: se não puder criar e/ou acessar a planilha.
        """
        logger.debug(f"Xlsx.get_sheet: xlsx_file={xlsx_file} sheet_name={sheet_name} create_sheet={create_sheet}")
        try:
            self.workbook = self._get_workbook(xlsx_file=xlsx_file, create_xlsx=create_sheet)
            if sheet_name not in self.workbook.sheetnames and create_sheet:
                self.workbook.create_sheet(sheet_name)
                worksheet = self.workbook[sheet_name]
                logger.debug(f"Xlsx.get_sheet: created sheet {sheet_name}")
                return worksheet
            elif sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                logger.debug(f"Xlsx.get_sheet: loaded sheet {sheet_name}")
                return worksheet
            else:
                raise ValueError(f'Planilha "{sheet_name}" não existe no arquivo "{xlsx_file}".')
        except Exception as error_x:
            logger.critical(f'Não foi possível retornar as planilhas do arquivo "{xlsx_file}".\nErro: {error_x}')
            raise RuntimeError(f'Não foi possível retornar as planilhas do arquivo "{xlsx_file}".\nErro: {error_x}') from error_x
    
    def get_cell_value(self, xlsx_file: str, sheet_name: str, cell_reference: str) -> str:
        """Retorna o valor da célula especificada.

        :param sheet_name: Nome da planilha que contém a célula.
        :param cell_reference: Referência da célula (ex: 'A1').
        :returns: valor da célula.
        :raises RuntimeError: se ocorrer qualquer falha.
        """
        logger.debug(f"Xlsx.get_cell_value: xlsx_file={xlsx_file} sheet_name={sheet_name} cell_reference={cell_reference}")
        try:
            self.workbook = self._get_workbook(xlsx_file=xlsx_file)
            sheet = self.get_sheet(xlsx_file=xlsx_file, sheet_name=sheet_name, create_sheet=False)
            value = sheet[cell_reference].value
            logger.debug(f"Xlsx.get_cell_value: value={value}")
            return value
        except Exception as error_x:
            logger.critical(f'Não foi possível obter o valor da célula "{cell_reference}" na planilha "{sheet_name}".\nErro: {error_x}')
            raise RuntimeError(f'Não foi possível obter o valor da célula "{cell_reference}" na planilha "{sheet_name}".\nErro: {error_x}') from error_x
    
    def set_cell_value(self, xlsx_file: str, sheet_name: str, cell_reference: str, value: any, create_sheet: bool = False, create_xlsx: bool = False) -> None:
        """Define o valor da célula especificada.

        :param sheet_name: nome da planilha que contém a célula.
        :param cell_reference: referência da célula (ex: 'A1').
        :param value: valor a ser definido na célula.
        :param create_sheet: criar planilha se não existir?
        :returns None: se não ocorrer nenhuma falha.
        :raises RunTimeError: se ocorrer qualquer falha.
        """
        logger.debug(f"Xlsx.set_cell_value: xlsx_file={xlsx_file} sheet_name={sheet_name} cell_reference={cell_reference} value={value} create_sheet={create_sheet} create_xlsx={create_xlsx}")
        try:
            self.workbook = self._get_workbook(xlsx_file=xlsx_file, create_xlsx=create_xlsx)
            sheet = self.get_sheet(xlsx_file=xlsx_file, sheet_name=sheet_name, create_sheet=create_sheet)
            sheet[cell_reference] = value
            logger.debug(f"Xlsx.set_cell_value: set {cell_reference} to {value}")
        except Exception as error_x:
            logger.critical(f'Não foi possível mudar o valor da célula "{cell_reference}" para "{value}".\nErro: {error_x}.')
            raise RuntimeError(f'Não foi possível mudar o valor da célula "{cell_reference}" para "{value}".\nErro: {error_x}.') from error_x

    def save_workbook(self, xlsx_file: str) -> None:
        """Função para salvar o arquivo self.workbook.
        
        :returns None: se não ocorrer nenhuma falha.
        :raises RunTimeError: se ocorrer alguma falha."""
        logger.debug(f"Xlsx.save_workbook: xlsx_file={xlsx_file}")
        try:
            self.workbook.save(filename=xlsx_file)
            logger.debug(f"Xlsx.save_workbook: saved {xlsx_file}")
        except Exception as error_x:
            logger.critical(f'Erro ao salvar o arquivo "{xlsx_file}".\nErro: {error_x}')
            self.printautomation.print_error()
            raise RuntimeError(f'Erro ao salvar o arquivo "{xlsx_file}"\nErro: {error_x}') from error_x


if __name__ == '__main__':
    xlsx_file = '.\\workbooks\\teste_01.xlsx'

    xlsx = Xlsx(process_id='0001', process_type='rpa', process_machine='COOP_0001')

    # Cria o arquivo (não existe ainda) e a planilha "Planilha1"
    sheet = xlsx.get_sheet(xlsx_file=xlsx_file, sheet_name='Planilha1', create_sheet=True)
    xlsx.save_workbook(xlsx_file=xlsx_file)
    logger.debug(f'SHEET_NAMES_1: {xlsx.get_sheet_names(xlsx_file=xlsx_file)}')

    # Escreve valores em algumas células
    xlsx.set_cell_value(xlsx_file=xlsx_file, sheet_name='Planilha1', cell_reference='A1', value='valor1')
    xlsx.set_cell_value(xlsx_file=xlsx_file, sheet_name='Planilha1', cell_reference='B1', value=123)
    xlsx.save_workbook(xlsx_file=xlsx_file)

    # Lê os valores gravados
    a1_value = xlsx.get_cell_value(xlsx_file=xlsx_file, sheet_name='Planilha1', cell_reference='A1')
    b1_value = xlsx.get_cell_value(xlsx_file=xlsx_file, sheet_name='Planilha1', cell_reference='B1')
    logger.debug(f'A1: {a1_value} | B1: {b1_value}')

    # Cria uma segunda planilha
    xlsx.set_cell_value(xlsx_file=xlsx_file, sheet_name='Planilha2', cell_reference='A1', value='outra_planilha', create_sheet=True)
    xlsx.save_workbook(xlsx_file=xlsx_file)
    logger.debug(f'SHEET_NAMES_2: {xlsx.get_sheet_names(xlsx_file=xlsx_file)}')

    # Tenta acessar planilha inexistente sem create_sheet -> deve lançar RuntimeError
    try:
        xlsx.get_sheet(xlsx_file=xlsx_file, sheet_name='NaoExiste', create_sheet=False)
    except RuntimeError as error_x:
        logger.error(f'ERRO_ESPERADO: {error_x}')